import re
from io import BytesIO
from datetime import time, datetime

import openpyxl
import pandas as pd
import streamlit as st
import xml.etree.ElementTree as ET
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


# =========================
# App configuration
# =========================
st.set_page_config(page_title="VO2 XML Extractor", layout="wide")
st.title("VO₂ XML Extractor")
st.write(
    "Upload up to 4 Excel 2003 SpreadsheetML XML files: AD1, AD2, ED1, ED2, "
    "plus optional Excel files for 5000m and lactate data."
)


# =========================
# Constants
# =========================
NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

EXPECTED_LABELS = ["AD1", "AD2", "ED1", "ED2"]
KEEP_COLUMNS = ["t", "Marker", "V'O2/kg", "HF"]
PLOT_COLUMNS = ["V'O2/kg", "HF"]


# =========================
# Parsing helpers
# =========================
def row_values(row: ET.Element) -> list:
    values = []
    current_col = 1

    for cell in row.findall("./ss:Cell", NS):
        index_attr = cell.attrib.get(f"{{{NS['ss']}}}Index")
        if index_attr is not None:
            target_col = int(index_attr)
            while current_col < target_col:
                values.append(None)
                current_col += 1

        data = cell.find("./ss:Data", NS)
        values.append(data.text if data is not None else None)
        current_col += 1

    return values


def find_worksheet(root: ET.Element, worksheet_name: str) -> ET.Element:
    for ws in root.findall(".//ss:Worksheet", NS):
        name = ws.attrib.get(f"{{{NS['ss']}}}Name")
        if name == worksheet_name:
            return ws
    raise ValueError(f'Worksheet "{worksheet_name}" not found.')


def find_measurement_header(parsed_rows: list[list]) -> tuple[int, list]:
    for i, vals in enumerate(parsed_rows):
        if vals and len(vals) >= 4 and vals[:4] == ["t", "Phase", "Marker", "V'O2"]:
            return i, vals
    raise ValueError("Measurement header row not found.")


def extract_participant_info(parsed_rows: list[list], header_idx: int) -> dict:
    info = {
        "participant_id": None,
        "first_name": None,
        "last_name": None,
        "participant_name": None,
        "weight_kg": None,
        "pre_weight": None,
        "post_weight": None,
        "pre_5000m": None,
        "post_5000m": None,
        "lactate_by_test": {},
    }

    search_rows = parsed_rows[:header_idx]

    for row in search_rows:
        if not row:
            continue

        row_clean = [str(x).strip() if x is not None else "" for x in row]

        for i, cell in enumerate(row_clean):
            cell_lower = cell.lower()

            if cell_lower == "id" and i + 1 < len(row_clean):
                value = row_clean[i + 1].strip()
                if value:
                    info["participant_id"] = value

            if cell_lower == "vorname" and i + 1 < len(row_clean):
                value = row_clean[i + 1].strip()
                if value:
                    info["first_name"] = value

            if cell_lower == "nachname" and i + 1 < len(row_clean):
                value = row_clean[i + 1].strip()
                if value:
                    info["last_name"] = value

            if cell_lower == "name" and i + 1 < len(row_clean):
                value = row_clean[i + 1].strip()
                if value and not info["participant_name"]:
                    info["participant_name"] = value

            if cell_lower.startswith("gewicht") and i + 1 < len(row_clean):
                value = row_clean[i + 1].strip().replace(",", ".")
                match = re.search(r"\d+(?:\.\d+)?", value)
                if match:
                    try:
                        info["weight_kg"] = float(match.group(0))
                    except Exception:
                        pass

    first = info.get("first_name")
    last = info.get("last_name")

    if first or last:
        info["participant_name"] = " ".join([x for x in [first, last] if x])

    return info


def extract_data_rows(parsed_rows: list[list], header_idx: int, header: list) -> list[list]:
    data_rows = []

    for vals in parsed_rows[header_idx + 2:]:
        if not vals or vals[0] in (None, ""):
            break

        if len(vals) < len(header):
            vals = vals + [None] * (len(header) - len(vals))

        data_rows.append(vals[:len(header)])

    if not data_rows:
        raise ValueError("No measurement rows found.")

    return data_rows


def parse_spreadsheetml(
    xml_bytes: bytes,
    worksheet_name: str = "MetasoftStudio"
) -> tuple[pd.DataFrame, dict]:
    root = ET.fromstring(xml_bytes)

    worksheet = find_worksheet(root, worksheet_name)

    table = worksheet.find("./ss:Table", NS)
    if table is None:
        raise ValueError("No table found in worksheet.")

    rows = table.findall("./ss:Row", NS)
    parsed_rows = [row_values(r) for r in rows]

    header_idx, header = find_measurement_header(parsed_rows)
    participant_info = extract_participant_info(parsed_rows, header_idx)
    data_rows = extract_data_rows(parsed_rows, header_idx, header)

    df = pd.DataFrame(data_rows, columns=header)
    return df, participant_info


# =========================
# Transformation helpers
# =========================
def time_to_seconds(series: pd.Series) -> pd.Series:
    td = pd.to_timedelta(
        series.astype(str).str.replace(",", ".", regex=False),
        errors="coerce"
    )
    return td.dt.total_seconds()


def keep_requested_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    present_columns = [c for c in KEEP_COLUMNS if c in df.columns]
    missing_columns = [c for c in KEEP_COLUMNS if c not in df.columns]

    if not present_columns:
        raise ValueError("None of the requested columns were found.")

    return df[present_columns].copy(), missing_columns


def trim_to_abbruch(df: pd.DataFrame) -> pd.DataFrame:
    if "Marker" not in df.columns:
        return df

    cutoff_idx = df[df["Marker"].astype(str).str.strip().str.lower() == "abbruch"].index
    if not cutoff_idx.empty:
        return df.loc[:cutoff_idx[0]]

    return df


def convert_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for col in ["V'O2/kg", "HF"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def add_time_seconds(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "t" in df.columns:
        df["t_seconds"] = time_to_seconds(df["t"])

    return df


def add_forward_rolling_averages(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "t_seconds" not in df.columns:
        return df

    df = df.sort_values("t_seconds").reset_index(drop=True)
    valid_time = df["t_seconds"].notna()

    if not valid_time.any():
        return df

    temp = df.loc[valid_time].copy()

    max_t = temp["t_seconds"].max()
    temp["t_rev"] = max_t - temp["t_seconds"]

    temp = temp.sort_values("t_rev")
    temp.index = pd.to_timedelta(temp["t_rev"], unit="s")

    for col in ["V'O2/kg", "HF"]:
        if col in temp.columns:
            temp[f"{col}_20s_fwd_ma"] = temp[col].rolling("20s").mean()
            temp[f"{col}_30s_fwd_ma"] = temp[col].rolling("30s").mean()

    temp = temp.sort_values("t_seconds")

    for col in ["V'O2/kg", "HF"]:
        for window in [20, 30]:
            ma_col = f"{col}_{window}s_fwd_ma"
            if ma_col in temp.columns:
                df[ma_col] = pd.NA
                df.loc[valid_time, ma_col] = temp[ma_col].values

    return df


# =========================
# Metric helpers
# =========================
def get_time_window_mask(df: pd.DataFrame, start_s: int, end_s: int) -> pd.Series:
    t = pd.to_numeric(df["t_seconds"], errors="coerce")
    return (t >= start_s) & (t <= end_s)


def calculate_time_to_abbruch(df: pd.DataFrame):
    if "t_seconds" not in df.columns:
        return pd.NA

    t_valid = pd.to_numeric(df["t_seconds"], errors="coerce")
    if t_valid.notna().any():
        return float(t_valid.max())

    return pd.NA


def calculate_max_vo2_20s(df: pd.DataFrame):
    col = "V'O2/kg_20s_fwd_ma"
    if col not in df.columns:
        return pd.NA

    value = pd.to_numeric(df[col], errors="coerce").max()
    return float(value) if pd.notna(value) else pd.NA


def calculate_max_vo2_30s(df: pd.DataFrame):
    col = "V'O2/kg_30s_fwd_ma"
    if col not in df.columns:
        return pd.NA

    value = pd.to_numeric(df[col], errors="coerce").max()
    return float(value) if pd.notna(value) else pd.NA


def calculate_mean_vo2_5_10(df: pd.DataFrame):
    if "t_seconds" not in df.columns or "V'O2/kg" not in df.columns:
        return pd.NA

    mask = get_time_window_mask(df, 300, 600)
    value = pd.to_numeric(df.loc[mask, "V'O2/kg"], errors="coerce").mean()
    return float(value) if pd.notna(value) else pd.NA


def calculate_hf_max(df: pd.DataFrame):
    if "HF" not in df.columns:
        return pd.NA

    value = pd.to_numeric(df["HF"], errors="coerce").max()
    return float(value) if pd.notna(value) else pd.NA


def calculate_hf_max_5_10(df: pd.DataFrame):
    if "t_seconds" not in df.columns or "HF" not in df.columns:
        return pd.NA

    mask = get_time_window_mask(df, 300, 600)
    value = pd.to_numeric(df.loc[mask, "HF"], errors="coerce").max()
    return float(value) if pd.notna(value) else pd.NA


def calculate_mean_hf_5_10(df: pd.DataFrame):
    if "t_seconds" not in df.columns or "HF" not in df.columns:
        return pd.NA

    mask = get_time_window_mask(df, 300, 600)
    value = pd.to_numeric(df.loc[mask, "HF"], errors="coerce").mean()
    return float(value) if pd.notna(value) else pd.NA


def calculate_metrics(df: pd.DataFrame, missing_columns: list[str]) -> dict:
    return {
        "max_vo2_20s": calculate_max_vo2_20s(df),
        "max_vo2_30s": calculate_max_vo2_30s(df),
        "mean_vo2_5_10": calculate_mean_vo2_5_10(df),
        "hf_max": calculate_hf_max(df),
        "hf_max_5_10": calculate_hf_max_5_10(df),
        "mean_hf_5_10": calculate_mean_hf_5_10(df),
        "missing_columns": missing_columns,
        "time_to_abbruch": calculate_time_to_abbruch(df),
    }


def pair_average(metrics_dict: dict, file_a: str, file_b: str, metric_name: str):
    values = []

    for key in [file_a, file_b]:
        value = metrics_dict.get(key, {}).get(metric_name, pd.NA)
        if pd.notna(value):
            values.append(float(value))

    return sum(values) / len(values) if values else pd.NA


# =========================
# Excel helpers: 5000m
# =========================
def excel_time_to_string(value):
    if value is None:
        return None

    if isinstance(value, time):
        if value.microsecond:
            hundredths = int(round(value.microsecond / 10000))
            return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}.{hundredths:02d}"
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"

    if isinstance(value, datetime):
        if value.microsecond:
            hundredths = int(round(value.microsecond / 10000))
            return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}.{hundredths:02d}"
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"

    return str(value).strip()


def load_5000m_lookup(excel_bytes: bytes, sheet_name: str = "150m & 5000m") -> dict:
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found in Excel file.')

    ws = wb[sheet_name]
    lookup = {}

    # Expected structure:
    # row 4+ data
    # B = ID, D = pre 5000m, G = post 5000m
    for row in range(4, ws.max_row + 1):
        participant_id = ws.cell(row, 2).value   # B
        pre_5000m = ws.cell(row, 4).value        # D
        post_5000m = ws.cell(row, 7).value       # G

        if participant_id is None:
            continue

        participant_id = str(participant_id).strip()
        if not participant_id:
            continue

        lookup[participant_id] = {
            "pre_5000m": excel_time_to_string(pre_5000m),
            "post_5000m": excel_time_to_string(post_5000m),
        }

    return lookup


def enrich_participant_with_5000m(participant_info: dict | None, lookup: dict | None) -> dict | None:
    if not participant_info:
        return participant_info

    participant_info = participant_info.copy()
    participant_info.setdefault("pre_5000m", None)
    participant_info.setdefault("post_5000m", None)

    if not lookup:
        return participant_info

    participant_id = participant_info.get("participant_id")
    if not participant_id:
        return participant_info

    match = lookup.get(str(participant_id).strip())
    if not match:
        return participant_info

    participant_info["pre_5000m"] = match.get("pre_5000m")
    participant_info["post_5000m"] = match.get("post_5000m")
    return participant_info


# =========================
# Excel helpers: lactate
# =========================
def load_lactate_lookup(excel_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
    lookup = {}

    def normalize_text(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def normalize_key(value) -> str:
        return normalize_text(value).lower()

    def try_float(value):
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        text = normalize_text(value).replace(",", ".")
        if not text:
            return None

        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None

        try:
            return float(match.group(0))
        except Exception:
            return None

    def is_speed_value(value) -> bool:
        num = try_float(value)
        return num is not None and 5 <= num <= 25

    def is_lactate_value(value) -> bool:
        num = try_float(value)
        return num is not None and 0.5 <= num <= 20

    def format_speed_label(value) -> str | None:
        num = try_float(value)
        if num is None:
            return None

        rounded = round(num, 2)
        if float(rounded).is_integer():
            return f"{int(rounded)} km/h"
        return f"{rounded:g} km/h"

    def find_participant_id(ws) -> str | None:
        title = normalize_text(ws.title)
        if re.fullmatch(r"Z\d{3}", title, flags=re.IGNORECASE):
            return title.upper()

        max_scan_rows = min(ws.max_row, 12)
        max_scan_cols = min(ws.max_column, 12)

        for row in range(1, max_scan_rows + 1):
            for col in range(1, max_scan_cols + 1):
                key = normalize_key(ws.cell(row, col).value)
                if key != "id":
                    continue

                for delta_col in [1, 2]:
                    candidate = normalize_text(ws.cell(row, col + delta_col).value)
                    if re.fullmatch(r"Z\d{3}", candidate, flags=re.IGNORECASE):
                        return candidate.upper()

                for delta_row in [1, 2]:
                    candidate = normalize_text(ws.cell(row + delta_row, col).value)
                    if re.fullmatch(r"Z\d{3}", candidate, flags=re.IGNORECASE):
                        return candidate.upper()

        return None

    def find_test_blocks(ws) -> list[tuple[str, int, int]]:
        blocks = []

        for row in range(1, min(ws.max_row, 12) + 1):
            for col in range(1, ws.max_column + 1):
                text = normalize_text(ws.cell(row, col).value).upper()
                if text in EXPECTED_LABELS:
                    blocks.append((text, row, col))

        return blocks

    def find_header_row(ws, test_row: int, label_col: int) -> int | None:
        for row in range(test_row + 1, min(ws.max_row, test_row + 30) + 1):
            texts = {
                offset: normalize_key(ws.cell(row, label_col + offset).value)
                for offset in range(-3, 3)
                if label_col + offset >= 1
            }

            if any("laktat" in text for text in texts.values()) and any(
                "km/h" in text or "kmh" in text for text in texts.values()
            ):
                return row

        return None

    def find_header_columns(ws, header_row: int, label_col: int) -> tuple[int | None, int | None, int | None]:
        speed_col = None
        lactate_col = None
        rpe_col = None

        for col in range(max(1, label_col - 4), min(ws.max_column, label_col + 3) + 1):
            text = normalize_key(ws.cell(header_row, col).value)

            if "km/h" in text or "kmh" in text:
                speed_col = col
            elif "laktat" in text:
                lactate_col = col
            elif "rpe" in text:
                rpe_col = col

        return speed_col, lactate_col, rpe_col

    def find_value_near_anchor(ws, row: int, anchor_col: int | None, validator, blocked_cols: set[int]) -> float | None:
        if anchor_col is None:
            return None

        for distance in [0, -1, 1, -2, 2]:
            col = anchor_col + distance
            if col < 1 or col > ws.max_column:
                continue
            if col in blocked_cols:
                continue

            value = ws.cell(row, col).value
            if validator(value):
                return try_float(value)

        return None

    def row_contains_stop_marker(ws, row: int, label_col: int) -> bool:
        for col in range(max(1, label_col - 4), min(ws.max_column, label_col + 1) + 1):
            text = normalize_key(ws.cell(row, col).value)
            if "rampe" in text or "abbruch" in text:
                return True

        return False

    def extract_test_values(ws, test_row: int, label_col: int) -> dict:
        header_row = find_header_row(ws, test_row, label_col)
        if header_row is None:
            return {}

        speed_col, lactate_col, rpe_col = find_header_columns(ws, header_row, label_col)
        blocked_for_speed = {col for col in [lactate_col, rpe_col] if col is not None}
        blocked_for_lactate = {col for col in [speed_col, rpe_col] if col is not None}
        values = {}
        pending_speed = None
        pending_row = None
        blank_rows = 0

        for row in range(header_row + 2, min(ws.max_row, header_row + 30) + 1):
            if row_contains_stop_marker(ws, row, label_col):
                break

            speed = find_value_near_anchor(
                ws,
                row,
                speed_col,
                is_speed_value,
                blocked_for_speed,
            )
            lactate = find_value_near_anchor(
                ws,
                row,
                lactate_col,
                is_lactate_value,
                blocked_for_lactate,
            )

            if speed is None and lactate is None:
                blank_rows += 1
                if blank_rows >= 4 and values:
                    break
                continue

            blank_rows = 0

            if speed is not None and lactate is not None:
                label = format_speed_label(speed)
                if label:
                    values[label] = lactate
                pending_speed = None
                pending_row = None
                continue

            if speed is not None:
                pending_speed = speed
                pending_row = row
                continue

            if lactate is not None and pending_speed is not None and pending_row is not None:
                if row - pending_row <= 2:
                    label = format_speed_label(pending_speed)
                    if label:
                        values[label] = lactate
                pending_speed = None
                pending_row = None

        return values

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        participant_id = find_participant_id(ws)
        if not participant_id:
            continue

        lookup.setdefault(participant_id, {})

        for test_label, test_row, label_col in find_test_blocks(ws):
            values = extract_test_values(ws, test_row, label_col)
            if values:
                lookup[participant_id].setdefault(test_label, {}).update(values)

    return lookup


def enrich_participant_with_lactate(
    participant_info: dict | None,
    lactate_lookup: dict | None
) -> dict | None:
    if not participant_info:
        return participant_info

    participant_info = participant_info.copy()
    participant_info.setdefault("lactate_by_test", {})

    if not lactate_lookup:
        return participant_info

    pid = participant_info.get("participant_id")
    if not pid:
        return participant_info

    participant_info["lactate_by_test"] = lactate_lookup.get(str(pid).strip(), {})
    return participant_info


def enrich_participant_with_weight_averages(
    participant_info: dict | None,
    participant_info_by_label: dict | None
) -> dict | None:
    if not participant_info:
        return participant_info

    participant_info = participant_info.copy()
    participant_info.setdefault("pre_weight", None)
    participant_info.setdefault("post_weight", None)

    if not participant_info_by_label:
        return participant_info

    def average_weight(labels: list[str]):
        values = []
        for label in labels:
            weight = participant_info_by_label.get(label, {}).get("weight_kg")
            if weight is not None and pd.notna(weight):
                values.append(float(weight))
        return sum(values) / len(values) if values else None

    participant_info["pre_weight"] = average_weight(["ED1", "ED2"])
    participant_info["post_weight"] = average_weight(["AD1", "AD2"])
    return participant_info


# =========================
# File processing
# =========================
def process_file(xml_bytes: bytes) -> tuple[pd.DataFrame, dict, dict]:
    df, participant_info = parse_spreadsheetml(xml_bytes)
    out, missing_columns = keep_requested_columns(df)

    out = trim_to_abbruch(out)
    out = convert_numeric_columns(out)
    out = add_time_seconds(out)
    out = add_forward_rolling_averages(out)

    metrics = calculate_metrics(out, missing_columns)
    return out, metrics, participant_info


# =========================
# Upload helpers
# =========================
def detect_label(filename: str):
    normalized = filename.lower()
    normalized = re.sub(r"[^a-z0-9]", "", normalized)

    for label in ["ad1", "ad2", "ed1", "ed2"]:
        if label in normalized:
            return label.upper()

    return None


def map_uploaded_files(uploaded_files) -> tuple[dict, list[str], list[tuple[str, str]]]:
    uploads = {}
    unmatched_files = []
    duplicate_labels = []

    for uploaded in uploaded_files:
        detected = detect_label(uploaded.name)

        if detected is None:
            unmatched_files.append(uploaded.name)
            continue

        if detected in uploads:
            duplicate_labels.append((detected, uploaded.name))
            continue

        uploads[detected] = uploaded

    return uploads, unmatched_files, duplicate_labels


def process_uploads(uploads: dict) -> tuple[dict, dict, dict | None, dict]:
    results = {}
    metrics_all = {}
    participant_info_first = None
    participant_info_by_label = {}

    for label, uploaded in uploads.items():
        try:
            raw = uploaded.read()
            out, metrics, participant_info = process_file(raw)

            results[label] = out
            metrics_all[label] = metrics
            participant_info_by_label[label] = participant_info

            if participant_info_first is None:
                participant_info_first = participant_info

        except Exception as e:
            st.error(f"{label} ({uploaded.name}): Parsing failed: {e}")

    return results, metrics_all, participant_info_first, participant_info_by_label


# =========================
# DataFrame builders
# =========================
def build_summary_df(metrics_all: dict, uploads: dict) -> pd.DataFrame:
    rows = []

    for label in EXPECTED_LABELS:
        if label in metrics_all:
            rows.append({
                "Upload": label,
                "Filename": uploads[label].name,
                "Max 30 s VO₂ average": metrics_all[label]["max_vo2_30s"],
                "Mean VO₂ from minute 5 to 10": metrics_all[label]["mean_vo2_5_10"],
                "HFmax": metrics_all[label]["hf_max"],
                "HFmax from minute 5 to 10": metrics_all[label]["hf_max_5_10"],
                "Time to abbruch (s)": metrics_all[label]["time_to_abbruch"],
            })

    return pd.DataFrame(rows)


def build_average_df(metrics_all: dict) -> pd.DataFrame:
    rows = [
        {
            "Category": "AD (AD1 + AD2)",
            "Average Max 30 s VO₂": pair_average(metrics_all, "AD1", "AD2", "max_vo2_30s"),
            "Average Mean VO₂ from minute 5 to 10": pair_average(metrics_all, "AD1", "AD2", "mean_vo2_5_10"),
            "Average HFmax": pair_average(metrics_all, "AD1", "AD2", "hf_max"),
            "Average HFmax from minute 5 to 10": pair_average(metrics_all, "AD1", "AD2", "hf_max_5_10"),
            "Average time to abbruch (s)": pair_average(metrics_all, "AD1", "AD2", "time_to_abbruch"),
        },
        {
            "Category": "ED (ED1 + ED2)",
            "Average Max 30 s VO₂": pair_average(metrics_all, "ED1", "ED2", "max_vo2_30s"),
            "Average Mean VO₂ from minute 5 to 10": pair_average(metrics_all, "ED1", "ED2", "mean_vo2_5_10"),
            "Average HFmax": pair_average(metrics_all, "ED1", "ED2", "hf_max"),
            "Average HFmax from minute 5 to 10": pair_average(metrics_all, "ED1", "ED2", "hf_max_5_10"),
            "Average time to abbruch (s)": pair_average(metrics_all, "ED1", "ED2", "time_to_abbruch"),
        },
    ]

    return pd.DataFrame(rows)


def speed_sort_value(speed_label):
    try:
        return float(str(speed_label).lower().replace(" km/h", "").replace(",", "."))
    except Exception:
        return 9999


def build_phase_average_lactate_df(
    lactate_by_test: dict | None,
    test_labels: list[str]
) -> pd.DataFrame:
    grouped_values = {}

    for test_label in test_labels:
        for speed, value in (lactate_by_test or {}).get(test_label, {}).items():
            num = pd.to_numeric(value, errors="coerce")
            if pd.notna(num):
                grouped_values.setdefault(speed, []).append(float(num))

    rows = []
    for speed, values in grouped_values.items():
        if values:
            rows.append({
                "Speed": speed,
                "Mean Lactate": sum(values) / len(values),
            })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(
            by="Speed",
            key=lambda col: col.map(speed_sort_value)
        ).reset_index(drop=True)

    return df


def format_metric_value(value, decimals: int = 1) -> str:
    if value is None or pd.isna(value):
        return "-"

    if isinstance(value, str):
        return value

    return f"{float(value):.{decimals}f}"


def format_integer_metric_value(value) -> str:
    if value is None or pd.isna(value):
        return "-"

    if isinstance(value, str):
        return value

    return str(int(round(float(value))))


def build_pre_post_summary_tables(metrics_all: dict, participant_info: dict | None) -> dict[str, pd.DataFrame]:
    participant_info = participant_info or {}

    overview_df = pd.DataFrame([
        {
            "Metric": "",
            "Pre": participant_info.get("pre_5000m") or "-",
            "Post": participant_info.get("post_5000m") or "-",
            "Label": "5000m Zeit (hh:mm:ss)",
        },
        {
            "Metric": "",
            "Pre": format_metric_value(participant_info.get("pre_weight"), 1),
            "Post": format_metric_value(participant_info.get("post_weight"), 1),
            "Label": "Körpergewicht (kg)",
        },
    ])

    oxygen_df = pd.DataFrame([
        {
            "Metric": "VO2max (ml/min/kg)",
            "Pre": format_metric_value(pair_average(metrics_all, "ED1", "ED2", "max_vo2_30s")),
            "Post": format_metric_value(pair_average(metrics_all, "AD1", "AD2", "max_vo2_30s")),
        },
        {
            "Metric": "VO2submax (ml/min/kg)",
            "Pre": format_metric_value(pair_average(metrics_all, "ED1", "ED2", "mean_vo2_5_10")),
            "Post": format_metric_value(pair_average(metrics_all, "AD1", "AD2", "mean_vo2_5_10")),
        },
    ])

    heart_df = pd.DataFrame([
        {
            "Metric": "HFmax (bpm)",
            "Pre": format_integer_metric_value(pair_average(metrics_all, "ED1", "ED2", "hf_max")),
            "Post": format_integer_metric_value(pair_average(metrics_all, "AD1", "AD2", "hf_max")),
        },
        {
            "Metric": "HFsubmax (bpm)",
            "Pre": format_integer_metric_value(pair_average(metrics_all, "ED1", "ED2", "mean_hf_5_10")),
            "Post": format_integer_metric_value(pair_average(metrics_all, "AD1", "AD2", "mean_hf_5_10")),
        },
    ])

    return {
        "Messwerte": overview_df,
        "Sauerstoffaufnahme": oxygen_df,
        "Herzfrequenz": heart_df,
    }


def create_lactate_chart_image(pre_df: pd.DataFrame, post_df: pd.DataFrame) -> BytesIO | None:
    if pre_df.empty and post_df.empty:
        return None

    fig, ax = plt.subplots(figsize=(8.6, 3.2))

    if not pre_df.empty:
        ax.plot(
            pre_df["Speed"].map(speed_sort_value),
            pre_df["Mean Lactate"],
            marker="o",
            linewidth=2.2,
            color="#1f77b4",
            label="Pre",
        )

    if not post_df.empty:
        ax.plot(
            post_df["Speed"].map(speed_sort_value),
            post_df["Mean Lactate"],
            marker="o",
            linewidth=2.2,
            color="#d62728",
            label="Post",
        )

    ax.set_title("Laktat Pre vs Post", fontsize=12, pad=8)
    ax.set_xlabel("Stufe (km/h)")
    ax.set_ylabel("Laktat (mmol/L)")
    ax.grid(True, linestyle="--", linewidth=0.6, alpha=0.45)
    ax.legend(frameon=False)

    image_bytes = BytesIO()
    fig.tight_layout()
    fig.savefig(image_bytes, format="png", dpi=200, bbox_inches="tight")
    plt.close(fig)
    image_bytes.seek(0)
    return image_bytes


def draw_placeholder_logo(pdf: canvas.Canvas, page_width: float, page_height: float):
    logo_x = 20 * mm
    logo_y = page_height - 28 * mm
    logo_w = page_width - 40 * mm
    logo_h = 14 * mm

    pdf.setFillColor(colors.HexColor("#e8eef7"))
    pdf.roundRect(logo_x, logo_y, logo_w, logo_h, 4 * mm, fill=1, stroke=0)
    pdf.setStrokeColor(colors.HexColor("#7d93b2"))
    pdf.roundRect(logo_x, logo_y, logo_w, logo_h, 4 * mm, fill=0, stroke=1)
    pdf.setFillColor(colors.HexColor("#39557a"))
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawCentredString(
        logo_x + logo_w / 2,
        logo_y + 4.8 * mm,
        "Individuelle Auswertung Zone 2 Studie",
    )


def draw_participant_box(pdf: canvas.Canvas, participant_info: dict | None, page_height: float):
    participant_info = participant_info or {}

    box_x = 20 * mm
    box_y = page_height - 52 * mm
    box_w = 95 * mm
    box_h = 16 * mm

    pdf.setFillColor(colors.white)
    pdf.setStrokeColor(colors.HexColor("#90a4bc"))
    pdf.roundRect(box_x, box_y, box_w, box_h, 3 * mm, fill=1, stroke=1)

    pdf.setFillColor(colors.HexColor("#34495e"))
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(box_x + 4 * mm, box_y + 10 * mm, "Name")
    pdf.drawString(box_x + 52 * mm, box_y + 10 * mm, "Teilnehmer ID")

    pdf.setFont("Helvetica", 11)
    pdf.drawString(
        box_x + 4 * mm,
        box_y + 4.2 * mm,
        str(participant_info.get("participant_name") or "-"),
    )
    pdf.drawString(
        box_x + 52 * mm,
        box_y + 4.2 * mm,
        str(participant_info.get("participant_id") or "-"),
    )


def draw_compact_info_table(
    pdf: canvas.Canvas,
    title: str,
    summary_df: pd.DataFrame,
    table_x: float,
    table_y_top: float,
    table_width: float,
    first_header: str = "Messwert",
):
    row_h = 7.5 * mm
    col_widths = [table_width * 0.52, table_width * 0.24, table_width * 0.24]
    headers = [first_header, "Pre", "Post"]

    if title:
        pdf.setFillColor(colors.HexColor("#34495e"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(table_x, table_y_top + 3.2 * mm, title)

    def draw_cell(x, y, w, h, text, fill_color=None, bold=False):
        if fill_color is not None:
            pdf.setFillColor(fill_color)
            pdf.rect(x, y, w, h, fill=1, stroke=0)

        pdf.setStrokeColor(colors.HexColor("#bcc9d6"))
        pdf.rect(x, y, w, h, fill=0, stroke=1)
        pdf.setFillColor(colors.HexColor("#1f2d3d"))
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 8.5)
        pdf.drawString(x + 2 * mm, y + 2.5 * mm, str(text))

    x = table_x
    y = table_y_top
    for header, width in zip(headers, col_widths):
        draw_cell(x, y, width, row_h, header, colors.HexColor("#dfe9f3"), bold=True)
        x += width

    current_y = table_y_top - row_h
    for _, row in summary_df.iterrows():
        x = table_x
        metric_text = row["Metric"]
        if "Label" in summary_df.columns and (not metric_text):
            metric_text = row["Label"]

        draw_cell(x, current_y, col_widths[0], row_h, metric_text)
        x += col_widths[0]
        draw_cell(x, current_y, col_widths[1], row_h, row["Pre"])
        x += col_widths[1]
        draw_cell(x, current_y, col_widths[2], row_h, row["Post"])
        current_y -= row_h

    return current_y - 4 * mm


def draw_summary_table(
    pdf: canvas.Canvas,
    title: str,
    summary_df: pd.DataFrame,
    table_x: float,
    table_y_top: float,
    first_header: str = "Messwert",
):
    row_h = 10 * mm
    col_widths = [78 * mm, 34 * mm, 34 * mm]
    headers = [first_header, "Pre", "Post"]

    if title:
        pdf.setFillColor(colors.HexColor("#34495e"))
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(table_x, table_y_top + 4 * mm, title)

    def draw_cell(x, y, w, h, text, fill_color=None, bold=False):
        if fill_color is not None:
            pdf.setFillColor(fill_color)
            pdf.rect(x, y, w, h, fill=1, stroke=0)

        pdf.setStrokeColor(colors.HexColor("#bcc9d6"))
        pdf.rect(x, y, w, h, fill=0, stroke=1)
        pdf.setFillColor(colors.HexColor("#1f2d3d"))
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 10)
        pdf.drawString(x + 3 * mm, y + 3.6 * mm, str(text))

    x = table_x
    y = table_y_top
    for header, width in zip(headers, col_widths):
        draw_cell(x, y, width, row_h, header, colors.HexColor("#dfe9f3"), bold=True)
        x += width

    current_y = table_y_top - row_h
    for _, row in summary_df.iterrows():
        x = table_x
        metric_text = row["Metric"]
        if "Label" in summary_df.columns and (not metric_text):
            metric_text = row["Label"]

        draw_cell(x, current_y, col_widths[0], row_h, metric_text)
        x += col_widths[0]
        draw_cell(x, current_y, col_widths[1], row_h, row["Pre"])
        x += col_widths[1]
        draw_cell(x, current_y, col_widths[2], row_h, row["Post"])
        current_y -= row_h

    return current_y - 6 * mm


def create_results_pdf(
    participant_info: dict | None,
    metrics_all: dict,
    pre_lactate_df: pd.DataFrame,
    post_lactate_df: pd.DataFrame,
) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    page_width, page_height = A4

    draw_placeholder_logo(pdf, page_width, page_height)
    draw_participant_box(pdf, participant_info, page_height)

    summary_tables = build_pre_post_summary_tables(metrics_all, participant_info)
    chart_bytes = create_lactate_chart_image(pre_lactate_df, post_lactate_df)
    table_x = 20 * mm
    table_top_y = page_height - 72 * mm
    table_width = page_width - 40 * mm

    current_y = draw_compact_info_table(
        pdf,
        "Messwerte",
        summary_tables["Messwerte"],
        table_x,
        table_top_y,
        table_width,
        first_header="",
    )
    current_y = draw_compact_info_table(
        pdf,
        "Sauerstoffaufnahme",
        summary_tables["Sauerstoffaufnahme"],
        table_x,
        current_y,
        table_width,
        first_header="VO2 (Sauerstoffaufnahme)",
    )
    current_y = draw_compact_info_table(
        pdf,
        "Herzfrequenz",
        summary_tables["Herzfrequenz"],
        table_x,
        current_y,
        table_width,
        first_header="HF (Herzfrequenz)",
    )

    if chart_bytes is not None:
        chart_x = 20 * mm
        chart_y = 16 * mm
        chart_w = page_width - 40 * mm
        chart_h = max(30 * mm, current_y - chart_y)
        pdf.drawImage(
            ImageReader(chart_bytes),
            chart_x,
            chart_y,
            width=chart_w,
            height=chart_h,
            preserveAspectRatio=True,
            mask="auto",
        )

    pdf.save()
    return buffer.getvalue()


def render_pdf_export_button(
    participant_info: dict | None,
    metrics_all: dict,
    pre_lactate_df: pd.DataFrame,
    post_lactate_df: pd.DataFrame,
):
    pdf_bytes = create_results_pdf(
        participant_info=participant_info,
        metrics_all=metrics_all,
        pre_lactate_df=pre_lactate_df,
        post_lactate_df=post_lactate_df,
    )
    participant_info = participant_info or {}
    participant_id = str(participant_info.get("participant_id") or "Unbekannt").strip()
    participant_name = str(participant_info.get("participant_name") or "Teilnehmer").strip()
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", participant_name).strip("_") or "Teilnehmer"
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "_", participant_id).strip("_") or "Unbekannt"
    file_name = f"{safe_id}_{safe_name}.pdf"

    st.download_button(
        label="PDF herunterladen",
        data=pdf_bytes,
        file_name=file_name,
        mime="application/pdf",
        key="download_results_pdf",
        use_container_width=False,
    )


# =========================
# UI helpers
# =========================
def render_upload_status(uploads: dict):
    st.subheader("Detected XML uploads")

    rows = []
    for label in EXPECTED_LABELS:
        rows.append({
            "Slot": label,
            "Status": "Uploaded" if label in uploads else "Missing",
            "Filename": uploads[label].name if label in uploads else ""
        })

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def render_upload_feedback(
    unmatched_files: list[str],
    duplicate_labels: list[tuple[str, str]],
    uploads: dict
):
    if unmatched_files:
        st.warning(
            "Could not detect AD1, AD2, ED1, or ED2 from filename: "
            + ", ".join(unmatched_files)
        )

    if duplicate_labels:
        st.error(
            "Duplicate uploads are not allowed. Ignored files: "
            + ", ".join([f"{filename} (detected as {label})" for label, filename in duplicate_labels])
        )

    missing_expected = [label for label in EXPECTED_LABELS if label not in uploads]
    if missing_expected:
        st.info("Missing uploads: " + ", ".join(missing_expected))


def render_participant_info(participant_info: dict | None):
    if not participant_info:
        return

    st.subheader("Participant information from first processed XML")

    col1, col2 = st.columns(2)
    with col1:
        st.metric("Participant ID", participant_info.get("participant_id") or "Not found")
    with col2:
        st.metric("Name", participant_info.get("participant_name") or "Not found")

    col3, col4 = st.columns(2)
    with col3:
        st.metric("5000m pre", participant_info.get("pre_5000m") or "Not found")
    with col4:
        st.metric("5000m post", participant_info.get("post_5000m") or "Not found")


def render_summary_section(summary_df: pd.DataFrame, avg_df: pd.DataFrame):
    st.subheader("Summary metrics per upload")
    st.dataframe(summary_df, use_container_width=True)

    st.subheader("Averaged values by category")
    st.dataframe(avg_df, use_container_width=True)


def render_chart(df: pd.DataFrame):
    plot_cols = [c for c in PLOT_COLUMNS if c in df.columns]

    if "t_seconds" not in df.columns or not plot_cols:
        return

    chart_df = (
        df[["t_seconds"] + plot_cols]
        .dropna(subset=["t_seconds"])
        .set_index("t_seconds")
    )

    if not chart_df.empty:
        st.line_chart(chart_df)


def render_download_button(label: str, df: pd.DataFrame):
    csv_bytes = df.to_csv(index=False).encode("utf-8")

    st.download_button(
        label=f"Download {label} CSV",
        data=csv_bytes,
        file_name=f"{label}_vo2_hf_timeseries.csv",
        mime="text/csv",
        key=f"download_{label}"
    )


def render_result_section(label: str, df: pd.DataFrame, metrics: dict, filename: str):
    st.subheader(f"{label} data ({filename})")
    st.dataframe(df, use_container_width=True)

    st.subheader(f"{label} time series")
    render_chart(df)

    if metrics["missing_columns"]:
        st.warning(
            f"{label}: Missing columns in this file: "
            + ", ".join(metrics["missing_columns"])
        )

    render_download_button(label, df)


def render_lactate_tables(participant_info):
    if not participant_info:
        return pd.DataFrame(), pd.DataFrame()

    lactate = participant_info.get("lactate_by_test", {})

    if not lactate:
        return pd.DataFrame(), pd.DataFrame()

    st.subheader("Lactate values per test")

    for test in EXPECTED_LABELS:
        values = lactate.get(test, {})
        if not values:
            continue

        st.markdown(f"**{test}**")

        df = pd.DataFrame([
            {"Speed": k, "Lactate": v}
            for k, v in values.items()
        ])

        if not df.empty:
            df = df.sort_values(
                by="Speed",
                key=lambda col: col.map(speed_sort_value)
            )

        st.dataframe(df, use_container_width=True, hide_index=True)

    st.subheader("Lactate averages by phase")

    pre_df = build_phase_average_lactate_df(lactate, ["ED1", "ED2"])
    post_df = build_phase_average_lactate_df(lactate, ["AD1", "AD2"])

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Pre (ED1 + ED2)**")
        if pre_df.empty:
            st.info("No pre lactate values found.")
        else:
            st.dataframe(pre_df, use_container_width=True, hide_index=True)

    with col2:
        st.markdown("**Post (AD1 + AD2)**")
        if post_df.empty:
            st.info("No post lactate values found.")
        else:
            st.dataframe(post_df, use_container_width=True, hide_index=True)

    return pre_df, post_df


# =========================
# Main app flow
# =========================
uploaded_files = st.file_uploader(
    "Upload XML files (AD1, AD2, ED1, ED2)",
    type=["xml"],
    accept_multiple_files=True
)

performance_file = st.file_uploader(
    "Upload Excel file with 150m & 5000m results",
    type=["xlsx"],
    key="performance_file"
)

lactate_file = st.file_uploader(
    "Upload lactate Excel file",
    type=["xlsx"],
    key="lactate_file"
)

if len(uploaded_files) > 4:
    st.error("Please upload no more than 4 XML files.")
    st.stop()

uploads, unmatched_files, duplicate_labels = map_uploaded_files(uploaded_files)

render_upload_status(uploads)
render_upload_feedback(unmatched_files, duplicate_labels, uploads)

performance_lookup = None
if performance_file is not None:
    try:
        performance_lookup = load_5000m_lookup(performance_file.read())
        st.success("Performance Excel loaded successfully.")
    except Exception as e:
        st.error(f"Performance Excel parsing failed: {e}")

lactate_lookup = None
if lactate_file is not None:
    try:
        lactate_lookup = load_lactate_lookup(lactate_file.read())
        st.success("Lactate file loaded.")
    except Exception as e:
        st.error(f"Lactate parsing failed: {e}")

results, metrics_all, participant_info_first, participant_info_by_label = process_uploads(uploads)

participant_info_first = enrich_participant_with_5000m(
    participant_info_first,
    performance_lookup
)

participant_info_first = enrich_participant_with_lactate(
    participant_info_first,
    lactate_lookup
)

participant_info_first = enrich_participant_with_weight_averages(
    participant_info_first,
    participant_info_by_label
)

if results:
    pre_lactate_df = build_phase_average_lactate_df(
        participant_info_first.get("lactate_by_test", {}) if participant_info_first else {},
        ["ED1", "ED2"],
    )
    post_lactate_df = build_phase_average_lactate_df(
        participant_info_first.get("lactate_by_test", {}) if participant_info_first else {},
        ["AD1", "AD2"],
    )

    render_pdf_export_button(
        participant_info=participant_info_first,
        metrics_all=metrics_all,
        pre_lactate_df=pre_lactate_df,
        post_lactate_df=post_lactate_df,
    )

    summary_df = build_summary_df(metrics_all, uploads)
    avg_df = build_average_df(metrics_all)

    render_participant_info(participant_info_first)
    pre_lactate_df, post_lactate_df = render_lactate_tables(participant_info_first)
    render_summary_section(summary_df, avg_df)

    for label in EXPECTED_LABELS:
        if label in results:
            render_result_section(
                label=label,
                df=results[label],
                metrics=metrics_all[label],
                filename=uploads[label].name
            )
else:
    st.info("Please upload at least one XML file.")
